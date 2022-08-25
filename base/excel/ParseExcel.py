#!/bin/env python
# coding:utf-8

import openpyxl
import os
import sys
import argparse
# python2 resolve encoding
# reload(sys)
# sys.setdefaultencoding("utf8")


class processExcel():
    def __init__(self, filepath, destpath):
        #self.file = filepath
        self.oriexcel = openpyxl.load_workbook(filename=filepath)
        self.orisheet = self.oriexcel.get_sheet_by_name(self.oriexcel.sheetnames[0])
        self.rows = self.orisheet.max_row
        self.columns = self.orisheet.max_column
        self.opts = ["销售信心[5, 6)", "销售动力[5.5, 6)", "理解与洞察[5, 6)", "抗压性[5.5, 7)", "建立联系[5.5, 6.5)", "维护客户关系[5.5, 6.5)", "成就[5.5, 6)", "成长机会[6, 8)", "8选3"]
        self.require = ["销售信心>=6", "销售动力>=6", "理解与洞察>=6","抗压性>=7", "建立联系>=6.5", "维护客户关系>=6.5", "成就[6, 8)", "激发购买意愿>=7", "8选1"]
        self.out = ["销售动力<5.5", "理解与洞察<5", "抗压性<5.5", "建立联系<5.5", "维护客户关系<5.5", "金钱回报<5", "6选3"]
        self.addColumn = ["通过", "淘汰"]
        self.fields = ["销售信心", "销售动力", "理解与洞察", "抗压性", "建立联系", "维护客户关系", "成就", "成长机会", "激发购买意愿", "金钱回报"]
        self.titles = list()
        self.destPath = destpath
        self.newwb = openpyxl.Workbook()
        self.res = "result"

    def read(self):
        print("sheet names : {}".format(self.oriexcel.sheetnames[0]))
        # 复制原有的数据
        self.copysheet(self.orisheet)
        self.appendTitle()
        # print("max rows :{}", ws1.max_row)
        # print("max columns:{}", ws1.max_column)
        # print(type(ws1.rows))
        # 分析每一行数据
        for ridx in range(2, self.rows+1):
            print("process idx :{}".format(ridx))
            date = self.getline(ridx)
            self.copyOridate(date, ridx)
            self.judgement_result(date, ridx)
        self.newwb.save(os.path.join(self.destPath, "2.xlsx"))

    # 对数据进行判断,是否满足可选   必选  淘汰,并根据结果填充对应的数据列
    def judgement_result(self, date, row):
        print("judgement date: {}".format(date))
        nwsheet = self.newwb.get_sheet_by_name(self.res)
        num_opts = 0
        num_require = 0
        num_out = 0
        for idx in range(0, len(date)):
            if idx == 0:
                if date[idx] >= 6:
                    num_require += 1
                elif 5 <= date[idx] < 6:
                   num_opts += 1
            if idx in range(1, 6):
                if 5.5 <= date[idx] < 6:
                    num_opts += 1
                elif date[idx] >= 6:
                    num_require += 1
                else:
                    num_out += 1
            if idx == 6:
                if 5.5 <= date[idx] < 6:
                    num_opts += 1
                elif 6 <= date[idx] < 8:
                    num_require += 1
            if idx == 7:
                if 6 <= date[idx] < 8:
                    num_opts += 1
            if idx == 8:
                if date[idx] >= 7:
                    num_require += 1
            if idx == 9:
                if date[idx] < 5:
                    num_out += 1
        print("opt_num :{}, require_num :{}, out_num:{}".format(num_opts, num_require, num_out))
        if num_opts >= 3:
            nwsheet.cell(row, self.columns+len(self.opts), "满足")
        else:
            nwsheet.cell(row, self.columns + len(self.opts), "不满足")
        if num_require >= 1:
            nwsheet.cell(row, self.columns+len(self.opts)+len(self.require), "满足")
        else:
            nwsheet.cell(row, self.columns + len(self.opts) + len(self.require), "不满足")
        if num_out >= 3:
            nwsheet.cell(row, self.columns+len(self.opts)+len(self.require)+len(self.out), "满足")
            nwsheet.cell(row, self.columns + len(self.opts) + len(self.require) + len(self.out)+2, "建议淘汰")
        else:
            nwsheet.cell(row, self.columns + len(self.opts) + len(self.require)+len(self.out), "不满足")

        if num_opts >= 3 and num_require >= 1:
            nwsheet.cell(row, self.columns+len(self.opts)+len(self.require)+len(self.out)+1, "通过")

        if not(num_opts >= 3 and num_require >=1) and num_out < 3:
            nwsheet.cell(row, self.columns + len(self.opts) + len(self.require) + len(self.out) + 1, "待定")



    # 拷贝源数据到对应的列
    def copyOridate(self, date, row):
        dsheet = self.newwb.get_sheet_by_name(self.res)
        outidx = 0
        requireidx = 1
        # print ("data: {}".format(date))
        for dd in range(0, len(date)):
            # print("dd = {}".format(dd))
            if dd == 0:
                dsheet.cell(row, self.columns+dd+1, date[dd])
                dsheet.cell(row, self.columns+len(self.opts)+1, date[dd])
            if dd in range(1, 6):
                dsheet.cell(row, self.columns + dd + 1, date[dd])
                dsheet.cell(row, self.columns + len(self.opts) + requireidx + 1, date[dd])
                dsheet.cell(row, self.columns+len(self.opts)+len(self.require) + outidx + 1, date[dd])
                requireidx += 1
                outidx += 1
            if dd == 6:
                dsheet.cell(row, self.columns + dd + 1, date[dd])
                dsheet.cell(row, self.columns + len(self.opts) + requireidx + 1, date[dd])
                requireidx += 1
            if dd == 7:
                dsheet.cell(row, self.columns + dd + 1, date[dd])
            if dd == 8:
                dsheet.cell(row, self.columns + len(self.opts) + requireidx + 1, date[dd])
                requireidx += 1
            if dd == 9:
                dsheet.cell(row, self.columns + len(self.opts) + len(self.require) + outidx + 1, date[dd])
                outidx += 1


    # 读取指定行的数据 row 读取哪一行的数据， count读取多少个数据
    def getline(self, row):
        dates = list()
        count = len(self.fields)
        for i in range(0, count):
            #print("field :{} index:{}".format(self.fields[i], self.titles.index(self.fields[i])+1))
            dates.append(self.orisheet.cell(row, self.titles.index(self.fields[i])+1).value)
        return dates

    def appendTitle(self):
        dsheet = self.newwb.get_sheet_by_name(self.res)
        count = len(self.opts) + len(self.require) + len(self.out)
        cols = dsheet.max_column
        dsheet.insert_cols(idx=(cols + count), amount=count)
        tiles = list()
        tiles.extend(self.opts)
        tiles.extend(self.require)
        tiles.extend(self.out)
        tiles.extend(self.addColumn)
        for idx in range(1, len(tiles)+1):
            dsheet.cell(1, cols+idx, tiles[idx-1])


    # 把原数据的sheet复制过来
    def copysheet(self, orisheet):
        dest = self.newwb.create_sheet(self.res, 0)
        rows = orisheet.max_row
        columns = orisheet.max_column
        # 提前插入行  插入列; 此操作会在idx之前插入行或者列
        dest.insert_rows(idx=1, amount=rows-1)
        dest.insert_cols(idx=1, amount=columns-1)
        # print("dest rows :{}", dest.max_row)
        # print("dest columns:{}", dest.max_column)
        for rw in range(1, rows+1):
            # print("rw = {}".format(rw))
            for col in range(1, columns+1):
                cel = orisheet.cell(rw, col)
                dest.cell(rw, col, cel.value)
                if rw == 1:
                    self.titles.append(cel.value)

def getParam():
    parse = argparse.ArgumentParser()
    parse.add_argument("-f", "--file", required=True, action='store', dest='filepath', help="The file to parse.")
    return parse.parse_args()


if __name__ == '__main__':
    #filePath = os.path.join(os.getcwd(), "1.xlsx")
    #print(filePath)
    parse = getParam()
    print("To process file :{}".format(parse.filepath))
    pros = processExcel(parse.filepath, os.getcwd())
    pros.read()

